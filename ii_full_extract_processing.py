#!/usr/bin/env python3
"""
ADF Pipeline Inventory
======================
Reads an Azure Data Factory JSON export and produces an Excel workbook
describing every Copy activity: what source feeds into which sink table/path,
with source system classification, expression resolution, and column mappings.

Usage:
    python adf_inventory.py [input.json] [output.xlsx]

Defaults:
    input  → adf_extract.json  (same directory)
    output → adf_inventory.xlsx
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Friendly source-system labels keyed by Linked Service name ────────────────
SOURCE_SYSTEM_MAP = {
    "LS_REST_SNV":             "ServiceNow",
    "LS_SP":                   "SharePoint",
    "LS_SP_Generic":           "SharePoint",
    "LS_WFM":                  "WFM SQL Server",
    "LS_MGMT_STUDIO":          "MGMT Studio SQL Server",
    "LS_ADLS_GEN2":            "ADLS Gen2",
    "ADLS_Storage_Gen2":       "ADLS Gen2",
    "DataLakeStore_Gen2":      "ADLS Gen2",
    "LS_ADLS_GEN2_Moog":       "ADLS Gen2",
    "LS_AzureDataStorage":     "ADLS Gen2",
    "LS_Blob_SaidExport":      "Azure Blob Storage",
    "Solarwinds_Http":         "SolarWinds",
    "LS_VManage":              "Cisco vManage",
    "LS_CISCO_vManage_HTTP":   "Cisco vManage",
    "LS_Cisco_VManage_REST":   "Cisco vManage",
    "LS_ATERNITY_HTTP_SERVER": "Aternity",
    "LS_Aternity_REST":        "Aternity",
    "Graylog_HTTP":            "Graylog",
    "GRAYLOG_HTTP_UPDATED":    "Graylog",
    "LS_REST_INFOBLOX":        "Infoblox",
    "LS_SEMS_REST_API":        "SEMS REST API",
    "LS_SPOG_COSMOS_DB":       "SPOG CosmosDB",
    "LS_CosmosDbNoSql":        "CosmosDB",
    "LS_CosmosDb_NoSql":       "CosmosDB",
    "LS_SPOG_PostgreSQL":      "SPOG PostgreSQL",
    "LS_SQL_DB":               "Azure SQL DB",
    "LS_PRESTO_SQL_DB":        "Azure SQL DB",
    "LS_M365":                 "Microsoft 365",
    "HttpNextgen_Anonymous":   "NextGen AMS",
    "amsnextgen":              "NextGen AMS",
    "NextgenAMS":              "NextGen AMS",
    "AMS_graph2":              "NextGen AMS",
    "Httpnextgen":             "NextGen AMS",
    "NextGen_AMS_SP":          "NextGen AMS",
    "LS_GMOTool_Thor":         "GMO Tool",
    "GMOTool_Ls":              "GMO Tool",
    "LS_ThousandEyes_REST":    "ThousandEyes",
    "LS_AlertMedia_HTTP_json": "AlertMedia",
    "Http_OPMI_Anonymous":     "OPMI HTTP",
    "GRAPH_API_HTTP_LS":       "MS Graph API",
    "LS_Http_Server":          "HTTP Server",
}

LS_TYPE_FALLBACK = {
    "RestService":       "REST API",
    "HttpServer":        "HTTP Server",
    "AzureSqlDatabase":  "Azure SQL DB",
    "SqlServer":         "SQL Server",
    "AzureBlobFS":       "ADLS Gen2",
    "AzureBlobStorage":  "Azure Blob Storage",
    "CosmosDb":          "CosmosDB",
    "AzureDatabricks":   "Databricks",
}

# ── Column definitions ─────────────────────────────────────────────────────────
COLUMNS = [
    ("Pipeline Name",              28),
    ("Pipeline Folder",            20),
    ("Activity Name",              30),
    ("Inside ForEach",             18),
    ("Source System",              22),
    ("Source Dataset Name",        28),
    ("Source Dataset Type",        18),
    ("Source Linked Service",      28),
    ("Source Linked Service Type", 22),
    ("Source Detail",              55),
    ("Source Act. Parameters",     55),
    ("Sink Dataset Name",          28),
    ("Sink Dataset Type",          18),
    ("Sink Linked Service",        28),
    ("Sink Table / Path  (Raw)",   50),
    ("Sink Table / Path (Resolved)",50),
    ("Sink Schema  (Raw)",         22),
    ("Sink Schema (Resolved)",     22),
    ("Sink Write Behavior",        18),
    ("Pre-Copy Script",            45),
    ("Column Mapping Count",        8),
    ("Column Mappings",            90),
]

# ── Style constants ────────────────────────────────────────────────────────────
HDR_BG    = "1F3864"
HDR_FG    = "FFFFFF"
ALT_BG    = "DCE6F1"
THIN_SIDE = Side(style="thin", color="AAAAAA")
BORDER    = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


# ══════════════════════════════════════════════════════════════════════════════
# Expression helpers
# ══════════════════════════════════════════════════════════════════════════════

def _raw_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return json.dumps(val)
    return str(val)


def _resolve(val, act_params: dict) -> tuple[str, str]:
    """
    Given a dataset property value (plain string or ADF Expression dict)
    and the activity-level parameters, return (raw_repr, resolved_repr).
    """
    if val is None:
        return "", ""

    if isinstance(val, dict):
        if val.get("type") == "Expression":
            raw = val["value"]
        else:
            raw = json.dumps(val)
    else:
        raw = str(val)

    resolved = _substitute(raw, act_params)
    return raw, resolved


def _substitute(expr: str, act_params: dict) -> str:
    """Replace @dataset().PARAM occurrences with values from act_params."""
    def replacer(m):
        key = m.group(1)
        # Case-insensitive key lookup
        for k, v in act_params.items():
            if k.lower() == key.lower():
                if isinstance(v, dict):
                    if v.get("type") == "Expression":
                        return f"[Runtime: {v['value']}]"
                    return json.dumps(v)
                return str(v)
        return m.group(0)

    result = re.sub(r"@dataset\(\)\.(\w+)", replacer, expr)
    # Tag any remaining ADF runtime expressions
    if "@" in result and result == expr:
        return f"[Runtime: {expr}]"
    return result


# ══════════════════════════════════════════════════════════════════════════════
# Dataset field extractors
# ══════════════════════════════════════════════════════════════════════════════

def _extract_sql_fields(props: dict, act_params: dict) -> tuple:
    """Extract (table_raw, table_res, schema_raw, schema_res) for SQL datasets."""
    table_raw,  table_res  = _resolve(props.get("table"),                         act_params)
    schema_raw, schema_res = _resolve(props.get("schema_type_properties_schema"),  act_params)
    return table_raw, table_res, schema_raw, schema_res


def _extract_file_path(props: dict, act_params: dict) -> tuple:
    """Build a composite path string for file-based datasets."""
    loc = props.get("location", {})
    folder    = loc.get("folder_path")
    fname     = loc.get("file_name")
    container = loc.get("container") or loc.get("file_system")
    rel_url   = loc.get("relative_url")          # Json/HttpFile

    parts_raw, parts_res = [], []

    if container:
        r, v = _resolve(container, act_params)
        parts_raw.append(f"[{r}]"); parts_res.append(f"[{v}]")
    if folder:
        r, v = _resolve(folder, act_params)
        parts_raw.append(r); parts_res.append(v)
    if fname:
        r, v = _resolve(fname, act_params)
        parts_raw.append(r); parts_res.append(v)
    if rel_url:
        r, v = _resolve(rel_url, act_params)
        parts_raw.append(f"URL:{r}"); parts_res.append(f"URL:{v}")

    sep = " / "
    return sep.join(parts_raw) or _raw_str(loc), sep.join(parts_res) or _raw_str(loc)


def get_sink_table_path(ds: dict, act_params: dict) -> tuple:
    """
    Returns (ls_ref, ds_type, table_raw, table_res, schema_raw, schema_res).
    """
    props  = ds.get("properties", {})
    dtype  = props.get("type", "Unknown")
    ls_ref = props.get("linked_service_name", {}).get("reference_name", "")

    if dtype in ("AzureSqlTable", "SqlServerTable"):
        t_raw, t_res, s_raw, s_res = _extract_sql_fields(props, act_params)
    elif dtype in ("Parquet", "Binary", "DelimitedText", "Json", "Excel", "Xml", "HttpFile"):
        t_raw, t_res = _extract_file_path(props, act_params)
        s_raw, s_res = "", ""
    elif dtype == "RestResource":
        t_raw, t_res = _resolve(props.get("relative_url"), act_params)
        s_raw, s_res = "", ""
    else:
        t_raw = t_res = s_raw = s_res = ""

    return ls_ref, dtype, t_raw, t_res, s_raw, s_res


def get_source_detail(src_ds: dict, act_params: dict, linked_services: dict) -> str:
    """Human-readable string summarising what the source is pointing at."""
    props   = src_ds.get("properties", {})
    dtype   = props.get("type", "")
    ls_ref  = props.get("linked_service_name", {}).get("reference_name", "")
    ls_props = linked_services.get(ls_ref, {}).get("properties", {})

    parts = []

    if dtype == "RestResource":
        _, v = _resolve(props.get("relative_url"), act_params)
        parts.append(f"Endpoint: {v}")
        base = ls_props.get("url", "")
        if base:
            parts.append(f"Base URL: {base}")

    elif dtype in ("AzureSqlTable", "SqlServerTable"):
        _, s = _resolve(props.get("schema_type_properties_schema"), act_params)
        _, t = _resolve(props.get("table"), act_params)
        if s:
            parts.append(f"Schema: {s}")
        if t:
            parts.append(f"Table: {t}")
        conn = ls_props.get("connection_string", ls_props.get("server", ""))
        if conn:
            parts.append(f"Connection: {str(conn)[:100]}")

    elif dtype in ("Parquet", "Binary", "DelimitedText", "Excel", "Xml"):
        loc = props.get("location", {})
        _, folder = _resolve(loc.get("folder_path"), act_params)
        _, fname  = _resolve(loc.get("file_name"),   act_params)
        if folder:
            parts.append(f"Path: {folder}")
        if fname:
            parts.append(f"File: {fname}")

    elif dtype == "Json":
        loc = props.get("location", {})
        _, rel = _resolve(loc.get("relative_url"), act_params)
        if rel:
            parts.append(f"URL: {rel}")
        base = ls_props.get("url", "")
        if base:
            parts.append(f"Base URL: {base}")

    elif dtype == "HttpFile":
        base = ls_props.get("url", "")
        if base:
            parts.append(f"Base URL: {base}")

    return " | ".join(parts) if parts else dtype


# ══════════════════════════════════════════════════════════════════════════════
# Source system classification
# ══════════════════════════════════════════════════════════════════════════════

def classify_source(ls_name: str, linked_services: dict) -> str:
    if ls_name in SOURCE_SYSTEM_MAP:
        return SOURCE_SYSTEM_MAP[ls_name]
    ls_type = linked_services.get(ls_name, {}).get("properties", {}).get("type", "")
    return LS_TYPE_FALLBACK.get(ls_type, f"{ls_type or 'Unknown'} ({ls_name})")


# ══════════════════════════════════════════════════════════════════════════════
# Column mapping extractor
# ══════════════════════════════════════════════════════════════════════════════

def extract_mappings(translator: dict) -> tuple[int, str]:
    mappings = (translator or {}).get("mappings", [])
    if not mappings:
        return 0, ""
    pairs = [
        f"{m.get('source', {}).get('name', '?')}→{m.get('sink', {}).get('name', '?')}"
        for m in mappings
    ]
    return len(pairs), "; ".join(pairs)


# ══════════════════════════════════════════════════════════════════════════════
# Activity traversal
# ══════════════════════════════════════════════════════════════════════════════

def collect_copy_activities(
    activities: list,
    pipeline_name: str,
    pipeline_folder: str,
    foreach_name: str = "",
) -> list[dict]:
    results = []
    for act in activities:
        atype = act.get("type", "")

        if atype == "Copy":
            results.append({
                "pipeline_name":   pipeline_name,
                "pipeline_folder": pipeline_folder,
                "foreach_name":    foreach_name,
                "activity":        act,
            })

        elif atype == "ForEach":
            results.extend(collect_copy_activities(
                act.get("activities", []),
                pipeline_name, pipeline_folder,
                foreach_name=act.get("name", "ForEach"),
            ))

        elif atype == "IfCondition":
            for branch in ("if_true_activities", "if_false_activities"):
                results.extend(collect_copy_activities(
                    act.get(branch, []),
                    pipeline_name, pipeline_folder, foreach_name,
                ))

        elif atype in ("Until", "Switch"):
            for inner in [act.get("activities", [])] + [
                c.get("activities", []) for c in act.get("cases", [])
            ]:
                results.extend(collect_copy_activities(
                    inner, pipeline_name, pipeline_folder, foreach_name
                ))

    return results


# ══════════════════════════════════════════════════════════════════════════════
# Row builder
# ══════════════════════════════════════════════════════════════════════════════

def build_row(entry: dict, datasets: dict, linked_services: dict) -> list:
    act             = entry["activity"]
    pipeline_name   = entry["pipeline_name"]
    pipeline_folder = entry["pipeline_folder"]
    foreach_name    = entry["foreach_name"]
    activity_name   = act.get("name", "")

    # ── Source side ────────────────────────────────────────────────────────────
    inputs      = act.get("inputs", [])
    src_ref     = inputs[0].get("reference_name", "") if inputs else ""
    src_params  = inputs[0].get("parameters", {})    if inputs else {}
    src_ds      = datasets.get(src_ref, {})
    src_props   = src_ds.get("properties", {})
    src_type    = src_props.get("type", "Unknown")
    src_ls_ref  = src_props.get("linked_service_name", {}).get("reference_name", "")
    src_ls_type = linked_services.get(src_ls_ref, {}).get("properties", {}).get("type", "")
    src_system  = classify_source(src_ls_ref, linked_services)
    src_detail  = get_source_detail(src_ds, src_params, linked_services)
    src_params_raw = json.dumps(src_params, ensure_ascii=False) if src_params else ""

    # ── Sink side ──────────────────────────────────────────────────────────────
    outputs    = act.get("outputs", [])
    snk_ref    = outputs[0].get("reference_name", "") if outputs else ""
    snk_params = outputs[0].get("parameters", {})    if outputs else {}
    snk_ds     = datasets.get(snk_ref, {})

    snk_ls_ref, snk_type, t_raw, t_res, s_raw, s_res = get_sink_table_path(
        snk_ds, snk_params
    )

    sink_act       = act.get("sink", {})
    write_behavior = sink_act.get("write_behavior") or sink_act.get("sql_writer_use_table_lock", "")
    pre_script     = sink_act.get("pre_copy_script", "")

    # ── Column mappings ────────────────────────────────────────────────────────
    map_count, map_str = extract_mappings(act.get("translator", {}))

    return [
        pipeline_name,
        pipeline_folder,
        activity_name,
        foreach_name,
        src_system,
        src_ref,
        src_type,
        src_ls_ref,
        src_ls_type,
        src_detail,
        src_params_raw,
        snk_ref,
        snk_type,
        snk_ls_ref,
        t_raw,
        t_res,
        s_raw,
        s_res,
        str(write_behavior) if write_behavior else "",
        pre_script or "",
        map_count,
        map_str,
    ]


# ══════════════════════════════════════════════════════════════════════════════
# Excel writing
# ══════════════════════════════════════════════════════════════════════════════

def _hdr_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=HDR_FG, size=10)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    return c


def write_data_sheet(wb: Workbook, sheet_name: str, rows: list) -> int:
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        _hdr_cell(ws, 1, col_idx, col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    alt_fill  = PatternFill("solid", fgColor=ALT_BG)
    cell_font = Font(name="Arial", size=9)
    cell_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, 1):
            # Ensure only Excel-safe scalars reach the cell
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            elif val is None:
                val = ""
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font      = cell_font
            c.alignment = cell_align
            c.border    = BORDER
            if fill:
                c.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    return len(rows)


def write_summary_sheet(wb: Workbook, summary: dict, factory_name: str):
    ws = wb.create_sheet("Summary", 0)

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"ADF Inventory  –  {factory_name}")
    title_cell.font = Font(name="Arial", bold=True, size=14, color=HDR_BG)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 30

    # Sub-header
    ws.cell(row=2, column=1, value=f"Total Copy Activities: {sum(summary.values())}")
    ws.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=10, color="555555")
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 18

    # Table header
    for col_idx, label in enumerate(["Source System", "Copy Activities", "Sheet"], 1):
        _hdr_cell(ws, 4, col_idx, label)

    alt_fill  = PatternFill("solid", fgColor=ALT_BG)
    cell_font = Font(name="Arial", size=10)

    for row_idx, (system, count) in enumerate(sorted(summary.items()), 5):
        fill = alt_fill if row_idx % 2 == 1 else None
        for col_idx, val in enumerate([system, count, system[:31]], 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font      = cell_font
            c.alignment = Alignment(horizontal="left" if col_idx != 2 else "center")
            c.border    = BORDER
            if fill:
                c.fill = fill

    # Total row
    total_row = 5 + len(summary)
    c1 = ws.cell(row=total_row, column=1, value="TOTAL")
    c2 = ws.cell(row=total_row, column=2, value=sum(summary.values()))
    for c in (c1, c2):
        c.font   = Font(name="Arial", bold=True, size=10)
        c.border = BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    json_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("_DATA_AND_OUTPUTS\shell-32-eun-adf-okjtarmveaftjxpeuzj_full_extract.json")
    out_path  = Path(sys.argv[2]) if len(sys.argv) > 2 else json_path.with_name("adf_inventory.xlsx")

    print(f"Reading: {json_path}")
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    factory_name    = data.get("factory_name", "Unknown Factory")
    pipelines       = data.get("pipelines",       {})
    datasets        = data.get("datasets",         {})
    linked_services = data.get("linked_services",  {})

    print(f"Factory : {factory_name}")
    print(f"Pipelines: {len(pipelines)}  |  Datasets: {len(datasets)}  |  Linked Services: {len(linked_services)}")

    # Collect every Copy activity across all pipelines
    all_entries = []
    for pname, pipeline in pipelines.items():
        folder = ""
        if isinstance(pipeline.get("folder"), dict):
            folder = pipeline["folder"].get("name", "")
        entries = collect_copy_activities(pipeline.get("activities", []), pname, folder)
        all_entries.extend(entries)

    print(f"Copy activities found: {len(all_entries)}")

    # Build rows and group by source system
    grouped: dict[str, list] = defaultdict(list)
    for entry in all_entries:
        row = build_row(entry, datasets, linked_services)
        grouped[row[4]].append(row)   # row[4] = Source System

    print(f"Source systems: {sorted(grouped.keys())}")

    # Write workbook
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    summary: dict[str, int] = {}
    for system in sorted(grouped.keys()):
        count = write_data_sheet(wb, system, grouped[system])
        summary[system] = count
        print(f"  Sheet '{system}': {count} rows")

    write_summary_sheet(wb, summary, factory_name)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\nSaved → {out_path}")


if __name__ == "__main__":
    main()
